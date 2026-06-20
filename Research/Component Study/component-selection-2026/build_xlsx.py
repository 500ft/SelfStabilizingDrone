#!/usr/bin/env python3
"""Build Component_Selection_2026.xlsx — one sheet per component category.

Data source: 5 clustered deep-research streams (A propulsion, B avionics,
C vision/sensing, D FPV, E support) run 2026-06-19 against primary/vendor
sources. confidence = confirmed (primary source) | inferred (estimated/derived).
This script is the durable, structured record of those findings.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

COLS = ["rank", "part_model", "key_specs", "mass_g", "price_usd",
        "pros", "cons", "primary_source", "confidence", "rationale"]
WIDTHS = [5, 26, 40, 8, 11, 40, 40, 30, 11, 44]

HEAD_FILL = PatternFill("solid", fgColor="1F3864")
HEAD_FONT = Font(bold=True, color="FFFFFF", size=11)
CONF_FILL = PatternFill("solid", fgColor="E2EFDA")   # confirmed -> green
INF_FILL = PatternFill("solid", fgColor="FCE4D6")    # inferred  -> orange
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP_TOP = Alignment(wrap_text=True, vertical="top")

wb = Workbook()

def add_sheet(title, rows, note=None):
    ws = wb.create_sheet(title=title[:31])
    r0 = 1
    if note:
        ws.cell(1, 1, note).font = Font(italic=True, color="555555")
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLS))
        ws.cell(1, 1).alignment = WRAP_TOP
        ws.row_dimensions[1].height = 30
        r0 = 2
    for c, name in enumerate(COLS, 1):
        cell = ws.cell(r0, c, name)
        cell.fill = HEAD_FILL; cell.font = HEAD_FONT; cell.border = BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for i, row in enumerate(rows, start=r0 + 1):
        for c, key in enumerate(COLS, 1):
            val = row.get(key, "")
            cell = ws.cell(i, c, val)
            cell.alignment = WRAP_TOP; cell.border = BORDER
            if key == "confidence":
                cell.fill = CONF_FILL if str(val).startswith("confirmed") else INF_FILL
    for c, w in enumerate(WIDTHS, 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = ws.cell(r0 + 1, 1)
    ws.auto_filter.ref = f"A{r0}:{get_column_letter(len(COLS))}{r0 + len(rows)}"
    return ws

def R(rank, part, specs, mass, price, pros, cons, src, conf, why):
    return dict(zip(COLS, [rank, part, specs, mass, price, pros, cons, src, conf, why]))

# ----------------------------------------------------------------------------
# 04 MOTORS (critical)
motors = [
 R(1,"Happymodel EX1103 11000KV","9N12P, 1.5mm shaft, 1-2S; 121.9g @9.2A,68W on Gemfan 2023R (2S WOT)","3.8","8-10",
   "ONLY motor w/ published 2S+2\" thrust curve; 4x=488g => T/W 2.44 @200g; proven; cheap","Mfr test (not independent); 9.2A/motor => 36.8A pack peak","happymodel.cn EX1103 KV11000","confirmed",
   "Only end-to-end documented 2S+2023 curve; clears envelope with real numbers. BASELINE."),
 R(2,"BetaFPV 1103 11000KV","1.5mm shaft, M1.6 on 8.5mm BC, 14mm, 2S, JST-1.25","3.3-3.8","9-11",
   "Cleanest documented mount geometry for printed frame; explicit 2S targeting","No thrust table published; thrust inferred from EX1103","betafpv.com 1103","confirmed/inferred",
   "Electrically same class; best mounting docs; infer thrust ~ EX1103."),
 R(3,"T-Motor F1103-II 11000KV","9N12P, step-shaft, NMB502Z bearing, 2-3S","~5.1","13-16",
   "Premium bearing + anti-disengage step shaft = durability for crash-prone guarded build","Heaviest (~+5g AUW); thrust table not accessible; priciest","getfpv.com t-motor-f1103","confirmed/inferred",
   "Durability pick; demoted on weight + unconfirmed thrust."),
 R(4,"iFlight XING-E Nano 1103 10000KV","1.5mm shaft, 2-3S, rated 1.2-2.5\" props","~3.6","10-13",
   "Lightest confirmed mass; explicitly 2\" on 2S; reputable","10000KV < baseline => slightly less thrust; no 2S+2\" table","amazon iFlight 1103","confirmed/inferred",
   "Strong light pick; 10000KV under baseline so below 11000KV trio."),
 R(5,"Flywoo NIN 1103 V2 10500KV","1.5mm shaft, 2S; 7650/10500KV options","~3.5","11-13",
   "Light; 10500KV near baseline; good build rep","Specs from search summary only (page 403); no thrust table","getfpv.com flywoo-nin-1103","inferred",
   "Viable near-baseline alt; mid-rank for unconfirmed specs."),
 R(6,"Happymodel EX1103 8000KV","9N12P, 1.5mm shaft, 2-3S","3.7","8-10",
   "Same proven body; better for 2.5\" / lower current / longer 2S flight","8000KV on 2\"/2S underspins => likely <100g/motor; better on 3S","happymodel.cn EX1103 6/8/12k","confirmed",
   "Completeness; KV too low for 2S+2\" T/W>=2 unless prop grows."),
 R(7,"RCInPower GTS V3 (no 1103)","GTS V3 line is 1002/1003-class, NOT 1103","-","-",
   "Excellent bearings/build IF a 1003 substitution acceptable","Not a 1103; off-spec","racedayquads RCInPower GTS V3","confirmed",
   "Recorded only to note RCInPower has no 1103 in this line."),
 R(8,"Emax TH1103 7000KV","1.5mm shaft, Tinyhawk-class","~3.5","7-9",
   "Cheap, available","7000KV = 1S/light-2S whoop motor; won't reach T/W 2 on 2S+2\"","racedayquads emax-th1103","confirmed",
   "Lowest; KV far under baseline, fails thrust target."),
]
# 05 PROPS (critical)
props = [
 R(1,"Gemfan Hurricane 2023-3","2.0\"D x 2.3\" pitch, 3-blade, PC, 1.5mm bore, 3-hole T-mount, 52.2mm disk","0.88","~3/set",
   "THE prop in the EX1103 121.9g data point => removes thrust-inference risk; correct 1.5mm bore","3-hole hub must match bell; PC less flexible in crash","flywoo/getfpv 2023-3","confirmed",
   "Specifying it collapses all thrust uncertainty. BASELINE."),
 R(2,"HQProp T2X2.3X3","2.0\"D x 2.3\" pitch, 3-blade, PC, 1.5mm T-mount, ~10mm hub","~0.75-0.85","~3/set",
   "Exact 2.3\" pitch baseline; HQ Durable PC tolerates guard strikes; lightest","No 1103+2S curve; exact 2.3 mass interpolated","hqprop.com T2X2X3","confirmed/inferred",
   "Closest direct match to 2.3\" baseline w/ durable PC."),
 R(3,"HQProp T2X2.5X3","2.0\"D x 2.5\" pitch, 3-blade, PC, 1.5mm, 9.8mm hub","0.89","~3/set",
   "Higher pitch => more top-end thrust toward 450g 'desirable' line","Higher pitch raises current => worsens pack-current problem","amazon HQ T2X2.5X3","confirmed",
   "'More thrust' option; trades directly against battery current headroom."),
 R(4,"Gemfan 2020 D51 (ducted)","2.0\"D ~1.9\" pitch, 1.5mm bore, for 1103-1105, ducted profile","0.91-1.00","~3/set",
   "Bore+rating match; suits SHROUDED/ducted guard; durable PC","Lower pitch => less thrust/RPM; heavier; built for ducts","pyrodrone Gemfan 2020 D51","confirmed",
   "Matched prop IF guards become full ducts."),
 R(5,"Gemfan 2023X3 (bench ref)","2.0x2.3 3-blade; bench 149.2g @5.9A on 1105/3S; 2.06 g/W","0.88","~3/set",
   "Independent bench data exists for the geometry","Bench was 1105/3S not 1103/2S; same SKU as #1","oscarliang 2\" prop shootout","confirmed (diff cond.)",
   "Closest independent bench; condition mismatch keeps below #1."),
 R(6,"Azure 2035X3","2.0x3.5\" pitch 3-blade; 115.8g @4.2A, 2.22 g/W (1105/3S)","~0.9","~4/set",
   "Highest measured efficiency => best flight time","3.5\" pitch too aggressive for 1103; out of envelope; untested 1103/2S","oscarliang 2\" shootout","confirmed (diff cond.)",
   "Efficiency leader on paper; pitch above 2.0-2.3\" envelope."),
 R(7,"Emax Avan 2\" 3-blade","2.0\" 3-blade; 154.4g @6.4A, 1.98 (1105/3S)","~0.9","~4/set",
   "Highest raw thrust in shootout","Out-of-envelope test cond.; 1103 bore/pitch unconfirmed","oscarliang 2\" shootout","confirmed (diff cond.)",
   "Thrust reference only; 1105/3S not a 1103/2S rec."),
 R(8,"BetaFPV Gemfan 40mm 4-blade","~40mm (sub-2\"), 4-blade, 1.5mm; paired w/ BetaFPV 1103","~0.5-0.7","~3/set",
   "Officially paired w/ BetaFPV 1103 11000KV; very light","Under 2\" => below envelope, lower thrust ceiling","betafpv.com 1103","confirmed",
   "Mfr-paired prop; sub-2\" outside target envelope."),
]
# 06 BATTERY (critical) -- current screen: 4-motor peak = 36.8A
battery = [
 R(1,"GNB/Gaoneng 2S 550mAh 100C HV (XT30)","7.6V, 550mAh, 100C cont/200C burst = 55A cont","29","13-15",
   "LARGEST current margin (55A vs 36.8A peak); covers prop/KV upgrades; light","HV 8.7V full charge needs HV-rated ESC; XT30 +~1g","gaoneng.shop","confirmed",
   "Only pack w/ large headroom over peak; sag protection + upgrade room."),
 R(2,"Tattu 650mAh 2S 75C (XT30)","7.4V, 650mAh, 75C = 48.75A cont","~38","11-13",
   "Biggest tank => longest flight; comfortable 48.75A margin; trusted","Heaviest (~38g) eats AUW; mass to verify","getfpv.com tattu-650-2s-75c","confirmed/inferred",
   "Endurance pick if AUW allows; below #1 only on weight."),
 R(3,"BetaFPV LAVA 2S 550mAh 75C (XT30)","7.6V LiHV, 550mAh, 75C = 41.25A cont, 4.18Wh, 67x15.5x13.5","29.5","9-11(2pk)",
   "Passes current screen ~12% margin; full datasheet; cheap; right size","Only ~12% headroom; burst C not published","betafpv.com LAVA 2S 550","confirmed",
   "Cheapest pack clearing the screen; safe default."),
 R(4,"GNB 2S 550mAh 80C (XT30)","7.4V, 550mAh, 80C/160C burst = 44A cont","~29","12-14",
   "Solid 44A margin + 160C burst; light","Spec from listing not datasheet","amazon GNB 550 2S","confirmed (listing)",
   "Good middle option; ties LAVA on size w/ more margin."),
 R(5,"CNHL MiniStar 2S 450mAh 70C (XT30U)","7.4V, 450mAh, 70C = 31.5A cont","~24-26","~4.5",
   "Lightest + cheapest IF you trim throttle ceiling","FAILS screen (31.5A<36.8A); burst C unpublished; mass unconf.","chinahobbyline MiniStar 450","confirmed (C-rating)",
   "Flags failure mode: lightest but ceiling under 4-motor peak."),
 R(6,"Auline 2S 450mAh 80C HV (XT30)","7.6V LiHV, 450mAh, 80C = 36A cont","28","8-10",
   "Confirmed 28g; best 450mAh option","36A ~ 36.8A peak => no margin / marginal fail; sag risk","getfpv.com auline-450-2s-80c","confirmed",
   "Strongest 450mAh but just meets peak w/ zero cushion; not primary."),
 R(7,"GNB 2S 450mAh 80C HV (XT30)","7.6V, 450mAh, 80C/160C = 36A cont/72A burst, 28g","28","10-11",
   "72A burst can absorb transient peak; full datasheet","Continuous 36A still under 36.8A","pyrodrone GNB 450 2S 80C","confirmed",
   "Better than Auline via 72A burst; continuous still borderline."),
 R(8,"BetaFPV 2S 550mAh 75C BT2.0","7.6V, 550mAh, 75C; BT2.0 connector","~28-30","9-11",
   "BT2.0 saves connector weight on micro","BT2.0 ~25-30A => connector is the limiter below 36.8A peak; use XT30","betafpv.com LAVA/BT2.0","confirmed/inferred",
   "Warns: at this power BT2.0 connector, not cell, limits. Spec XT30."),
]
# 01 FLIGHT CONTROLLER (critical)
fc = [
 R(1,"Holybro Kakute H7 Mini","H743 480MHz; ICM-42688-P; SPA06 baro; 6 UART; 5V/2A BEC; 20x20","5.5","~65",
   "ArduPilot+PX4 both OFFICIAL; 6 UART for MAVLink companion; baro for alt-hold","Pricey; overkill on Betaflight; 5V/2A BEC modest","holybro.com; docs.px4.io","confirmed",
   "Only 20x20/5.5g board first-class on BOTH ArduPilot+PX4 — what autonomous recovery needs. TOP for autonomy."),
 R(2,"Matek H743-Mini v3","H743; dual gyro ICM42688P+42605; DPS310; 5.5 UART; 20x20","~7","~60",
   "Most mature ArduPilot; dual-gyro redundancy","PX4 experimental; heavier","mateksys.com; ardupilot.org","confirmed",
   "Near-equal to #1; dual gyro is a real robustness win."),
 R(3,"Matek H743-Slim v3","H743; dual gyro; DPS310; 7 UART; 2-8S; 30.5x30.5","7","~65",
   "Most UARTs (7) for companion+GPS+ELRS+telem; robust ArduPilot","30.5x30.5 footprint large for a 2\"","getfpv.com matek-h743-slim","confirmed",
   "Most companion-friendly; only the footprint holds it back."),
 R(4,"Matek F405-MiniTE","F405 168MHz; ICM42605; OSD; 6 UART; 20x20/30x30","5","~35",
   "Lightest board that STILL runs ArduPilot; cheap","F4 limited headroom; no onboard baro","mateksys; getfpv.com","confirmed",
   "Lightest ArduPilot-capable; accept F4 limits + ext baro."),
 R(5,"SpeedyBee F405 Mini","F405; ICM42688P; baro+OSD; 6 UART; BT; 20x20","9.6","~30",
   "Cheap, BT tuning, baro, INAV/ArduPilot","Heaviest 20x20 here","speedybee.com; getfpv.com","confirmed",
   "Best value F4 w/ baro; 9.6g hurts on sub-200g."),
 R(6,"Flywoo GOKU GN405 Nano","F405; ICM42688P; baro; 6 UART; 16x16; TCXO ELRS option","~4-5","~40",
   "Tiny, integrated ELRS option, light","16x16 std; Betaflight/INAV only, NO ArduPilot","flywoo.net GN405","inferred",
   "Great for Betaflight stabilize-only; not an autonomy platform."),
 R(7,"BetaFPV Toothpick F4 1-2S 12A AIO ELRS","F411; 12A 4-in-1 BLHeli_S; ELRS option; 3 UART; 1-2S","~5","~40",
   "True AIO (FC+ESC+ELRS) ~5g; ideal 2S/1103; minimal wiring","F411 Betaflight-only; 3 UART tight","wrekd.com; betafpv.com","confirmed",
   "Mass-optimal single board for Betaflight self-level; rules out autonomy."),
 R(8,"Happymodel X12 AIO 5-in-1 ELRS","F4; 12A 4-in-1; OpenVTX; ELRS onboard; 1-2S; 30x30","5.1","~40",
   "All-in-one incl VTX+ELRS at 5.1g; lowest wiring/mass","F4 SPI Betaflight-only; few UART","happymodel.cn X12; getfpv.com","confirmed",
   "Most integrated/lightest (FC+ESC+VTX+RX). No autonomy headroom. TOP for Betaflight build."),
 R(9,"SpeedyBee F405 Mini stack FC","F405; ICM42688P; baro; 4-6 UART; 20x20","~6-9","~30",
   "Cheap pre-matched 20x20 stack; INAV/ArduPilot","Heavier; ArduPilot less common target","getfpv.com speedybee-stack","inferred",
   "Convenient stack; conceptually redundant with #5."),
 R(10,"Matek F405-STD","F405; baro on some; 5-6 UART; ArduPilot MatekF405","~6","~30",
   "Proven ArduPilot F4 lineage; flexible I/O","Older/heavier than MiniTE","ardupilot.org common-matekf405","inferred",
   "Completeness; MiniTE supersedes for this build."),
]
# 02 ESC / AIO (critical)
esc = [
 R(1,"Integrated 12A AIO (BetaFPV/Happymodel X12)","1-2S, 12A 4-in-1, BLHeli_S/Bluejay (on the FC)","(in FC)","(in FC)",
   "Right current class for 2S/1103; ZERO ESC wiring; lowest system mass","Fixed to its FC; BLHeli_S (no telem on some)","betafpv.com; happymodel.cn X12","confirmed",
   "ESC mass disappears into FC => system-mass winner for Betaflight build."),
 R(2,"HGLRC Forward FD13A 4-in-1","2-4S, 13A cont/20A burst, BLHeli_S, 16x16","3.2","~20",
   "Tiny/light; current matched to 2S/1103; cheap; DShot600","16x16 footprint; no BEC; BLHeli_S only","getfpv.com hglrc-fd13a","confirmed",
   "Mass-optimal STANDALONE if FC/ESC separated (e.g. w/ H7 FC)."),
 R(3,"HGLRC 28A 2-4S 4-in-1 20x20","2-4S, 28A cont/35A burst, BLHeli_S, 20x20 (M2)","4.2","~22",
   "True 20x20 to match H7/F4 FCs; light","Mild over-spec current","getfpv.com hglrc 20x20","confirmed",
   "Best 20x20-matched standalone for a Kakute H7/Matek H743 stack."),
 R(4,"Holybro Tekko32 F4 Mini 45A AM32","2-6S, 45A, AM32 firmware, 20x20","(stack)","~35",
   "Native Kakute H7 stack partner; AM32 telemetry; solid","Over-spec for 2S/1103; adds mass; pricey","holybro.com tekko32-mini-45a","confirmed",
   "Convenient Holybro stack but 45A overkill here."),
 R(5,"Holybro Tekko32 F4 Mini 50A AM32","2-6S, 50A, AM32, 20x20","15.6","~40",
   "AM32, bulletproof","15.6g FAR too heavy; current+mass both wrong","holybro.com tekko32-50a","confirmed",
   "DO NOT use on this build — wrong by a wide margin."),
 R(6,"SpeedyBee BLS 35A Mini V2 20x20","3-6S, 35A/45A burst, BLHeli_S, current sensor, 20x20","7.2","~25",
   "Cheap, current sensor, proven, 20x20","3-6S min (verify 2S); over-spec","speedybee.com BLS35A","confirmed",
   "Better for 3-4S 2.5\"; verify 2S idle behavior."),
 R(7,"Aikon AK32PIN 35A 6S 20x20","2-6S, 35A/45A, BLHeli_32, current sensor","10","~30",
   "BLHeli_32 full telemetry/programmable; 2S-capable","10g + over-spec","getfpv.com ak32pin","confirmed",
   "If you specifically want BLHeli_32 telem on 20x20."),
 R(8,"HGLRC 13A 2-3S 4-in-1 (BB2)","2-3S, 13A/20A burst, BLHeli_S, EFM8BB2, 20x20","3","~18",
   "Lightest standalone; exact current class; cheap","2-3S only (fine for 2S); no BEC/telem","rotorev hglrc-13a-bb2","confirmed",
   "Tied-lightest standalone, perfectly matched to 2S/1103."),
 R(9,"T-Motor F15A/F25A 2-4S 4-in-1","2-4S, 15-25A, BLHeli_S, 20x20","~4-6","25-30",
   "Quality FETs; light; right class","Availability varies; BLHeli_S only","T-Motor vendor pages","inferred",
   "Clean light option in the right band; verify SKU specs."),
 R(10,"HGLRC Forward 45A BLHeli_32 2-6S 20x20","2-6S, 45A, BLHeli_32, 20x20","~6-7","~30",
   "BLHeli_32, 2S-capable, 20x20","Over-spec current; heavier","getfpv.com hglrc-45a-bl32","confirmed",
   "Capable but over-spec; only if BLHeli_32 + HGLRC ecosystem wanted."),
]
# 06b VISION / COMPUTE (critical)
vision = [
 R(1,"Arduino Nicla Vision","STM32H747 M7@480/M4@240, 1MB RAM, GC2145 2MP cam, onboard IMU(LSM6DSOX)+VL53L1 ToF, no NPU","19.8","70-115",
   "All-in-one: cam+IMU+ToF on 20g — exactly the fusion for release detection; TFLite Micro/Edge Impulse; 22.9x22.9","No HW AI accel; 1MB RAM caps model; GC2145 rolling shutter","docs.arduino.cc ABX00051","confirmed",
   "Lightest board carrying its own motion sensors; release detection IS motion fusion. TOP."),
 R(2,"OpenMV Cam H7 Plus","M7@480; 32MB SDRAM+1MB SRAM; OV5640 5MP; MicroPython; TFLite; no NPU","17","80-100",
   "32MB SDRAM => bigger frames/models; mature CV; CAN/UART/MAVLink-friendly; 140-240mA","No HW accel; OV5640 rolling shutter; needs separate IMU","openmv.io","confirmed",
   "Best pure-vision MCU at weight; SDRAM is the differentiator. Lacks onboard IMU vs Nicla."),
 R(3,"OpenMV Cam RT1062","i.MX RT1062 M7@600; 32MB SDRAM; OV5640; WiFi/BT5.1; ~40FPS QVGA","~17","80-100",
   "Fastest OpenMV (600MHz); onboard WiFi/BT telemetry; rich FC I/O","No NPU; rolling shutter; weight not vendor-confirmed","sparkfun/openmv.io","inferred",
   "Higher clock + onboard radio; weight inferred from shared form factor."),
 R(4,"Seeed XIAO ESP32-S3 Sense","ESP32-S3 @240MHz; 8MB PSRAM; OV2640 2MP (detachable); mic; SD slot","~3","14-20",
   "Cheapest+lightest; 8MB PSRAM; ESP-DL/Edge Impulse; WiFi/BLE; 21x17.5","240MHz limits FPS/model; OV2640 rolling shutter (EOL); no IMU","wiki.seeedstudio.com","confirmed",
   "Mass+cost leader; runs tiny CNNs but weakest throughput. Best low-budget fallback."),
 R(5,"Google Coral Dev Board Micro","i.MX RT1176 M7@1GHz+M4; 64MB RAM; Edge TPU 4 TOPS; HM01B0 324x324 mono","unconf.","80-100",
   "REAL HW NPU (4 TOPS @2 TOPS/W) => genuine onboard CNN fast; 1GHz M7","Cam 324x324 mono only; weight UNCONFIRMED; rigid TFLite-EdgeTPU toolchain","coral.ai datasheet","confirmed",
   "Only sub-board-class true TPU => strongest real-time inference. CONFIRM MASS before designing in."),
 R(6,"Espressif ESP32-S3-EYE","ESP32-S3@240MHz; 8MB PSRAM; OV2640 2MP; LCD; mic","unconf.","50-60",
   "Reference ESP-WHO AI board; PSRAM; well-documented","LCD/sub-board add weight; 240MHz; rolling shutter; no IMU","espboards.dev; adafruit","confirmed",
   "Same class as XIAO but heavier/pricier; XIAO Sense is the better drone pick."),
 R(7,"Sipeed Maix (K210)","Dual RISC-V@400MHz + KPU ~1 TOPS; 8MB SRAM; VGA cam @30FPS","light","25-40",
   "Real NPU ~1 TOPS at low cost/weight; MaixPy CV","Aging toolchain; 8MB caps models; rolling shutter; quirky","maixduino.sipeed.com","confirmed",
   "Cheap NPU that genuinely runs onboard inference; weaker tooling than Coral."),
 R(8,"Sipeed Maix-II / CanMV-K230","Kendryte K230 dual C908 RISC-V; KPU ~13.7x K210; 512MB RAM; 3x MIPI-CSI","heavier","50-70",
   "Much stronger NPU than K210; MIPI-CSI; RVV1.0","Linux-class weight/power; likely over budget; immature FC integration","youyeetoo CanMV-K230","inferred",
   "Real horsepower but trends to SBC weight/power; marginal for 200g."),
 R(9,"Raspberry Pi Zero 2 W","Quad A53@1GHz; 512MB; CSI-2 cam; WiFi/BT; 1.2W typ/3W max","10","15-18",
   "Light (10g); full Linux+OpenCV+MAVLink (mavros); takes GLOBAL-shutter OV9281 via CSI","No NPU (CPU inference slow); +SD+cam+regulator grows system mass","datasheets.raspberrypi.com","confirmed",
   "Flexible Linux route; pair w/ global-shutter cam. CPU-only CNN is the catch."),
 R(10,"Luxonis OAK-D-Lite","Myriad X VPU 4 TOPS; OV7251 mono x2 GLOBAL shutter + IMX214 13MP; BMI270 IMU; USB","61","~169",
   "Powerful VPU; stereo depth; GLOBAL-shutter stereo; onboard IMU; full CNNs","61g + up to 5W blows weight AND power on 2S; needs USB host","shop.luxonis.com","confirmed",
   "Capability ceiling but disqualified on weight/power. Listed for completeness."),
]
# 07 CAMERA
camera = [
 R(1,"OV9281 (global shutter)","1MP 1280x800 mono, GLOBAL shutter, up to 60FPS@1MP, RAW8/10, MIPI CSI-2","~5-8","25-40",
   "GLOBAL SHUTTER => no motion blur/jello on a thrown drone; high FPS; CSI direct to Pi","Mono only; needs CSI host (Pi Zero 2W), not MCU DVP","ovt.com; arducam.com","confirmed",
   "Correct sensor for fast-motion release detection — kills rolling-shutter distortion."),
 R(2,"OV7251 (global shutter)","0.31MP VGA mono, GLOBAL shutter, up to 120FPS, MIPI CSI-2","~3-5","20-35",
   "Global shutter; ultra-small; high FPS (used in OAK-D-Lite)","Low res (VGA); mono; CSI host needed","leopardimaging; arducam","confirmed",
   "Lighter/lower-res global-shutter alt; fine for optical-flow / event detection."),
 R(3,"HM01B0 (Himax)","324x324 mono, very low power; bundled on Coral Dev Board Micro","<2","bundled",
   "Extremely low power; bundled w/ Coral TPU; fine for fast triggers","Very low res; mono","coral.ai datasheet","confirmed",
   "Free with Coral; resolution sufficient for binary release-event classification."),
 R(4,"OV5640","5MP 2592x1944, ROLLING shutter, 30FPS, DVP/MIPI","~3-5","10-20",
   "High res; bundled on OpenMV H7 Plus/RT1062; color","ROLLING shutter => blur/skew during fast release","openmv.io; arducam","confirmed",
   "Default on OpenMV; usable but rolling shutter degrades the high-dynamics moment."),
 R(5,"GC2145","2MP color, rolling shutter; bundled on Nicla Vision","<2","bundled",
   "Bundled w/ Nicla Vision (#1 board); low mass; color","ROLLING shutter; modest res","docs.arduino.cc","confirmed",
   "Comes with top board; OK if release detection leans on the IMU."),
 R(6,"OV2640","UXGA 2MP, rolling shutter, onboard JPEG, DVP/SCCB","<2","5-10",
   "Cheap; JPEG offloads MCU; bundled on XIAO/ESP32","ROLLING shutter; being discontinued (->OV3660)","arducam.com","confirmed",
   "Default ESP32-class sensor; rolling shutter + EOL are the cons."),
 R(7,"OV7670","VGA 640x480, rolling shutter, 30FPS, DVP","<2","3-6",
   "Cheapest; widely supported","Low res; rolling shutter; no JPEG; dated","arducam.com","confirmed",
   "Legacy/budget only; no advantage over OV2640 here."),
]
# 11a OPTICAL FLOW
oflow = [
 R(1,"Matek 3901-L0X","PMW3901 flow + VL53L0X lidar, 8cm-200cm, UART MSP, 40mA, 36x12mm, 4.5-5.5V","2","22-30",
   "Flow + height in one 2g board; native INAV & ArduPilot 4.1+; single UART","VL53L0X height only to 2m; MSP not Betaflight-standard","mateksys; ardupilot.org","confirmed",
   "Indoor alt-hold + position-hold in one light part. DEFAULT for sub-250g indoor."),
 R(2,"PMW3901 (standalone)","optical flow only, 80mm-inf, UART 19200, 95Hz, 2.0-3.6V, 10mA, 14x11x5","0.6","10-20",
   "Lightest (0.6g); supported PX4/ArduPilot/INAV","No height sensor — needs separate rangefinder","docs.px4.io PMW3901","confirmed",
   "Lightest flow if you already have a rangefinder."),
 R(3,"ARK Flow","PAW3902 flow + AFBR-S50 30m + BMI088 IMU, DroneCAN","heavier","120-150",
   "30m range; better low-light/high-rate; onboard IMU; DroneCAN","Heaviest/priciest; needs CAN; over budget","docs.px4.io ark_flow","confirmed",
   "Technically superior but DroneCAN + mass = overkill here."),
 R(4,"PX4Flow","CMOS flow + sonar/lidar, I2C/MAVLink, dedicated MCU","~15","40-100",
   "Mature, well-documented; onboard processing","Heavy (~15g), large, legacy; over budget","docs.px4.io optical_flow","confirmed",
   "Classic but heavy/old; inappropriate for 200g."),
 R(5,"Cheerson/clone PMW3901","PMW3901-based clones, UART/SPI","~1-2","8-15",
   "Very cheap","Variable QA; thin docs","vendor/px4 gitbook","inferred",
   "Budget clone; only if cost-critical and you can validate."),
]
# 11b GPS
gps = [
 R(1,"Matek M10Q-5883","u-blox SAM-M10Q, 4 GNSS, 5-10Hz, QMC5883L compass, 20x20","8","30-40",
   "Latest M10; compass included; light; 20x20; ArduPilot/INAV","Indoor useless; +8g + mast","mateksys; defiancerc","confirmed",
   "Best modern light GPS+compass; lightest path to outdoor nav w/ heading."),
 R(2,"Holybro Micro M10 GPS","u-blox M10, 4 GNSS, up to 25Hz, IST8310 compass","14-16","30-45",
   "High update rate; quality IST8310 compass; PX4/ArduPilot","Heavier than Matek; case adds mass","docs.holybro.com","confirmed",
   "Excellent PX4 fit w/ strong compass; ~2x Matek mass."),
 R(3,"Beitian BN-220","u-blox M8, GPS+GLONASS, ~10Hz, NO compass, 22x20x6","5.3","12-18",
   "Lightest usable GPS; cheap; ArduPilot-documented","No compass (needs FC mag); older M8","ardupilot.org; rotorgeeks","confirmed",
   "Lightest; fine where FC has a mag and weight is critical."),
 R(4,"Holybro M9N Micro","u-blox M9N, 4 GNSS, up to 25Hz, IST8310 compass","~14-16","40-55",
   "Solid M9N; compass","Heavier; pricier; superseded by M10","docs.holybro.com","inferred",
   "M10 variants newer for similar mass; weight inferred."),
 R(5,"Beitian BN-880","u-blox M8, GPS+GLONASS, HMC5883L compass, 28x28x10","10-12","18-25",
   "Has compass; cheap","M8-era; large 28x28; heavier","ardupilot.org; BN-880 datasheet","confirmed",
   "Compass-equipped budget M8; large + older than M10."),
]
# 08 ELRS RX
elrs = [
 R(1,"BetaFPV ELRS Lite RX","SX1280, ESP8285, flat ceramic SMD ant, no PA/LNA, ~300m@25mW/1km@50mW","0.46","~12",
   "Lightest viable; flat ant easy to mount in a guard","No PA/LNA => lower telem power/sensitivity; single ant","betafpv.com elrs-lite","confirmed",
   "Effectively free mass; LOS range rarely the limit on a micro. DEFAULT."),
 R(2,"Happymodel EP2","SX1280, ESP8285, TCXO, SMD ceramic ant, 500Hz, 10x10x6","0.44","~10",
   "Lightest; TCXO thermal stability; integrated ant","Ceramic ant directional/lower gain; no diversity","happymodel.cn EP1/EP2","confirmed",
   "Tied-lightest; TCXO is a genuine link-lock plus."),
 R(3,"Happymodel EP1","SX1280, TCXO, omni dipole T ant (40/90mm), 500Hz","0.42 (no ant)","~10",
   "Omni dipole = best/most consistent pattern for agile micro; TCXO","External antenna to route/secure","happymodel.cn EP1","confirmed",
   "Lightest omni-dipole option for a tumbling/maneuvering craft."),
 R(4,"BetaFPV ELRS Nano RX","SX1280 + PA/LNA (100mW telem), dipole T ant, 500Hz","~1.8 (w/ant)","~17",
   "PA/LNA = best range+telem of single-ant set; strong link margin","~1.8g; heavier than Lite","betafpv.com elrs-nano","confirmed",
   "Step up for max link margin / telemetry for a companion feedback loop."),
 R(5,"RadioMaster RP1 V2","SX1280, ESP8285, UFL ant socket (full-range), WiFi/WebUI","2.2 (w/ant)","~13",
   "Swappable UFL antenna; WiFi config; full-range","UFL + ant add mass/bulk","radiomasterrc.com RP1","confirmed",
   "When you want to choose/upgrade the antenna."),
 R(6,"RadioMaster RP2 V2","SX1280, ESP8285, built-in ceramic ant, WiFi","0.55","~13",
   "Sub-gram, integrated ant, WiFi config","Ceramic lower gain; no diversity","pyrodrone RadioMaster RP2","confirmed",
   "0.55g integrated-ant alt to Lite/EP2 in RadioMaster ecosystem."),
 R(7,"RadioMaster RP3 V2 (diversity)","SX1280 + PA/LNA 100mW, antenna DIVERSITY 2x65mm T, 500/1000Hz","4.6 (w/ant)","~20",
   "Antenna diversity = far fewer dropouts when frame/guard shadows an ant; PA/LNA","4.6g heavy; two antennas to route","radiomasterrc.com RP3","confirmed",
   "Reliability play for a guarded frame that can shadow antennas; mass is the price."),
]
# 09 ANTENNAS (video) + control note
antennas = [
 R(1,"Lumenier Micro Dipole 5.8G (U.FL)","~2.5dBic, U.FL; analog + HDZero","1.0","~6",
   "Lightest mast; U.FL matches most nano VTX; broadband","Dipole (less gain than lollipop)","getfpv.com lumenier-micro-dipole-ufl","confirmed",
   "Best mass/perf for analog/HDZero video on a micro. (Control-link antenna comes with the ELRS RX.)"),
 R(2,"HDZero Whoop Lite stock dipole","5.8G dipole bundled w/ Whoop Lite VTX","~0.4","included",
   "Tuned to the VTX; near-zero mass","Fragile; low gain; HDZero band only","getfpv.com hdzero-whoop-lite","confirmed",
   "If going HDZero digital, the stock 0.4g dipole is lightest."),
 R(3,"Foxeer Micro Lollipop 5.8G (U.FL)","2.5dBi RHCP, U.FL, ~65mm","~1.5 (U.FL); 3.1 (MMCX)","~8/pr",
   "Circular pol rejects multipath; durable","Heavier; pick U.FL not the 3.1g MMCX","pyrodrone foxeer-micro-lollipop","confirmed/inferred",
   "CP for cleaner video in cluttered LOS; +~0.5g."),
 R(4,"Walksnail Avatar stock antenna (IPEX-1)","proprietary-tuned, IPEX-1","0.5","included",
   "Matched to Walksnail VTX; ultralight","Walksnail-only; IPEX-1 not interchangeable w/ MMCX","oscarliang Walksnail Mini","confirmed",
   "Use the bundled antenna for Walksnail digital."),
 R(5,"Lumenier Micro Dipole 5.8G (MMCX)","same as #1 but MMCX","~1","~7",
   "For VTXs w/ MMCX sockets","Most nano VTX use U.FL","getfpv.com lumenier-micro-dipole-mmcx","confirmed/inferred",
   "Only if your chosen VTX has an MMCX socket."),
]
# 10a FPV ANALOG
fpv_analog = [
 R(1,"Caddx Ant Lite + Foxeer Reaper Nano V2","Cam 1200TVL (3.7-18V) + 5.8G VTX 25/100/200/350mW (5V, 340mA@350mW)","~2.6","~35",
   "Lightest credible cam+VTX; VTX wide power; 350mW real LOS range","VTX is 5V-only => needs 5V BEC (NOT raw 2S)","pyrodrone (both parts)","confirmed",
   "Absolute min mass; both independently sourced. VTX needs a 5V rail."),
 R(2,"BetaFPV C04 + M04 module","1200TVL 160°, VTX 25-400mW, 5V, smart-audio","4.14 (w/ant)","~25",
   "One-piece cam+VTX+antenna, solder-free, 400mW","5V supply (needs BEC); heavier than discrete nano","betafpv.com C04; amazon","confirmed",
   "Lowest-risk integration; single-part simplicity."),
 R(3,"RunCam Nano 3 + Reaper Nano V2","Cam 800TVL 2.3mm 140° (1.1g) + 5V VTX","~1.9","~40",
   "Among the lightest cameras (1.1g), good low light","800TVL (lower); cam Vin unconfirmed","oscarliang RunCam Nano 3","confirmed/inferred",
   "Lightest camera; verify cam Vin tolerates 5V rail."),
 R(4,"Happymodel OVX306 OpenVTX + Caddx Ant Lite","VTX 37ch up to 400mW OpenVTX + 1200TVL cam","~3.0","~33",
   "OpenVTX (open firmware, BF VTX tables); 400mW; cheap","VTX mass inferred; 5V-class; whoop form","pyrodrone OVX306","inferred",
   "Good open-firmware option; verify VTX mass."),
 R(5,"Foxeer Razer Nano + Reaper Nano V2","Cam 1200TVL 1.8mm (4.5-25V, native 2S) + VTX","~4.7","~42",
   "Camera runs DIRECTLY off 2S (no cam BEC); strong low light","Camera heavy (3.9g) for a 'nano'","getfpv.com foxeer-razer-nano","confirmed",
   "Only option whose camera natively eats 2S; pay in grams."),
 R(6,"BetaFPV C03 + A03 400mW VTX","Cam 1200TVL 160° (3-5.5V) + VTX 25-400mW","~3+","~35",
   "Solderless JST-0.8 cam; 1200TVL; documented ecosystem","Cam ceiling 5.5V => needs 5V BEC; A03 mass unconf.","betafpv.com C03/A03","confirmed/inferred",
   "Solid plug-and-play; watch the 5.5V cam ceiling."),
 R(7,"RunCam Nano 4 + nano VTX","1200/800TVL 2.1mm 155° FOV cam + VTX","2.9 (cam)+VTX","~45",
   "Wide 155° FOV; waterproof variant; good night vision","Heavier than Nano 3, no res gain","shop.runcam.com Nano 4","confirmed",
   "Only over Nano 3 if you want the wider FOV."),
]
# 10b FPV DIGITAL
fpv_digital = [
 R(1,"HDZero Whoop Lite VTX + Nano Lite cam","2.8-13V (NATIVE 2S), ~200mW, ~7W draw, ~14ms frame latency","~6","~120",
   "Lightest production digital AND native 2S input; low latency","~7W draw heavy on 2S (cuts flight time/thermal); 200mW range","docs.hd-zero.com; getfpv.com","confirmed",
   "ONLY digital system fitting both mass AND 2S without a BEC. The realistic digital choice."),
 R(2,"Walksnail Avatar HD Mini 1S Lite Kit","cam 1.8g+VTX 5.1g+ant 0.5g; 3.1-5V; ~4W; 350mW; 22ms","8.7 (kit)","~130",
   "Lightweight, well-measured, good image, integrated kit","Input 3.1-5V ONLY => cannot run raw 2S (needs 5V BEC); ~4W","oscarliang Walksnail Mini 1S Lite","confirmed",
   "Mass-viable but 5V-input means no direct 2S."),
 R(3,"Walksnail Avatar HD Nano Kit V3","3.1-13V (native 2S), 500mW, 19-40ms; cam<3g+VTX~13g","~21 (kit)","~140",
   "Native 2S input; 500mW range; race mode 19-27ms","21g over the whole FPV budget alone; heavy metal VTX","caddxfpv.com Nano V3","confirmed",
   "Native-2S but too heavy; for completeness, not recommended."),
 R(4,"DJI O4 Air Unit (Lite)","3.7-13.2V (native 2S), 1/2\" CMOS, 20-35ms; 8.2-9.2g body","8.2-9.2 (unit)","~160",
   "Best image/sensor; light body; native 2S","Multi-W draw; needs DJI goggles only; mass climbs w/ antennas","dji.com; oscarliang O4 Lite","confirmed",
   "Body mass fine; power budget + goggle lock-in are the blockers."),
 R(5,"DJI O4 Air Unit Pro","native 2S input; 4K120 onboard recording","'>10 (unit)","~200",
   "Top image quality; onboard 4K recording","Heaviest/most power-hungry; outside sub-200g micro budget","getfpv.com o4-air-unit-pro","confirmed",
   "Not feasible for this airframe; closes out the DJI line honestly."),
]
# 12 BUZZER + LED
buzzer = [
 R(1,"VIFLY Finder 2 (V2)","Self-powered, 80mAh, 106 dB, push-button, ~6h","4.9","18-20",
   "Survives battery ejection; very loud; button control","4.9g notable on <200g build","getfpv.com vifly-finder-2","confirmed",
   "Best loudness self-powered finder; works after LiPo separation."),
 R(2,"VIFLY Finder Mini","Self-powered, 40mAh, 90 dB","2.8","~15",
   "Half the mass of V2; still self-powered","16 dB quieter; smaller battery","oscarliang drone-buzzers","confirmed",
   "Best mass/function trade for sub-200g. RECOMMENDED."),
 R(3,"JHE42B_S","FC-driven 5V, 110 dB, built-in status LED, 75mAh buffer","4.3","~6",
   "Loudest; cheap; integrated status LED","Needs FC 5V/signal; not fully self-powered","oscarliang beeper","confirmed",
   "Value pick if you want loud + LED and have an FC beeper pad."),
 R(4,"HellGate Buzzer","Self-powered, 40mAh, 88 dB, ~2-week standby","2.1","~13",
   "Lightest self-powered; long standby","Quietest of the set","oscarliang drone-buzzers","confirmed",
   "Absolute min-mass self-powered choice."),
 R(5,"WS2812 LED strip (10-20 LED, LANTIANRC/Malockos)","5V addressable RGB, 0.23A@20 LED, 1.0mm 3P","0.4-0.8","4-8(4pk)",
   "Negligible mass; BF/INAV status/arming/orientation patterns","Needs FC LED pad + 5V","amazon Malockos; newegg","confirmed",
   "Standard status/orientation LED, near-zero mass. Pair w/ a buzzer."),
]
# 13 WIRING / CONNECTORS
wiring = [
 R(1,"XT30 (battery connector)","~30A continuous; lowest measured sag in 2S test","1.6","~0.5/pr",
   "Lowest resistance; 20A peak easily; huge battery selection","Heaviest micro option (still ~1.6g)","oscarliang micro-connectors","confirmed",
   "At 10-20A peak low resistance beats 1g savings. RECOMMENDED battery connector."),
 R(2,"BT2.0 (battery connector)","9A rated; near-XT30 to ~12A","0.6","~0.4/pr",
   "Light; low sag at moderate current; BetaFPV ecosystem","9A rating marginal at 20A peak","getfpv.com BT2.0","confirmed",
   "Only if peak stays <=~12A; under-rated for 20A bursts."),
 R(3,"22 AWG silicone (battery leads)","~7A bundle class, fits 2S micro currents","N/A","~3/m",
   "Flexible, low mass for the current; standard for 2\"","20 AWG safer if sustained >15A","brushlesswhoop","inferred",
   "22 AWG main leads + 26 AWG motor leads = the 2S/2\" scheme."),
 R(4,"JST-SH 1.0mm (signal)","1.0mm pitch, low-current signal (RX/GPS/I2C)","N/A","~2/pack",
   "Tiny, keyed, standard for micro RX/peripherals","Fragile; not for power","brushlesswhoop","confirmed",
   "Signal-side standard for RX/GPS/LED breakouts."),
 R(5,"MMCX (VTX antenna)","RF coax connector, repeatable mate","N/A","~2 ea",
   "Antenna pops off on crash vs snapping the pad","RF loss if dirty/loose","getfpv.com","inferred",
   "Standard for swappable VTX antenna on micro."),
]
# 14 CHARGER
charger = [
 R(1,"ToolkitRC M6AC","AC 100W / DC 300W, 15A, 1-6S, dual input, balance, cell checker","N/A","~60",
   "AC+DC in one (no separate PSU); covers 2S w/ huge margin","Pricier than DC-only","toolkitrc.com M6AC","confirmed",
   "Single-box AC charger; no extra power supply to buy. RECOMMENDED."),
 R(2,"ToolkitRC M7AC","AC 100W / DC 300W, 15A, 2-6S, XT30+XT60","N/A","~75",
   "More headroom; dual output ports","Overkill for 2S micro","getfpv.com M7AC","confirmed",
   "If you may grow to bigger packs later."),
 R(3,"ToolkitRC M6 V2 (DC)","DC 150W, 10A, 1-6S, servo/cell checker","N/A","~42",
   "Cheapest capable; multifunction tester","Needs separate DC PSU","getfpv.com M6 V2","confirmed",
   "Best value if you already own a 12-24V supply."),
 R(4,"ToolkitRC M4AC","AC 30W, 2.5A, 1-4S, XT30, compact","N/A","~30",
   "Cheapest AC; amperage matched to small 2S packs","Low power; one pack at a time","toolkitrc.com M4AC","confirmed",
   "Sized exactly for 2S micro charging."),
 R(5,"BetaFPV BT2.0/PH2.0 charger + parallel board","1S-cell boards / 6-port, ~1A","N/A","10-20",
   "Cheap parallel charging of whoop packs","1S-cell oriented; NOT a 2S-series balance solution","betafpv.com BT2.0 charger","confirmed",
   "Only relevant if you also run 1S whoop packs."),
]
# 16 FASTENERS / STANDOFFS
fasteners = [
 R(1,"M2 steel socket/button-head kit","M2 4-12mm; T6/T8 hex","N/A","~10",
   "Won't strip/break in crash; correct for motor mounting","Heavier than nylon","oscarliang screws; pyrodrone M2","confirmed",
   "Motor + primary structural screws should be steel (w/ blue threadlocker)."),
 R(2,"M2/M2.5 nylon standoff+screw assortment (Readytosky 150pc)","M2/M2.5/M3 nylon hex standoffs, screws, nuts","N/A","~10",
   "Lightest for FC-stack standoffs; non-conductive; won't rust","Snaps in hard crash","amazon Readytosky","confirmed",
   "Nylon standoffs save mass + isolate the FC stack."),
 R(3,"M2 soft-mount silicone grommets","Anti-vibration grommets for FC/motor mounts","N/A","~5/set",
   "Decouples gyro from frame vibration => cleaner gyro data","Over-tightening kills damping","oscarliang soft-mount","confirmed",
   "CRITICAL for a self-stabilizing build — protects gyro signal. Don't over-torque."),
 R(4,"M2.5 nylon FC standoffs","M2.5 nylon, common in micro FC stacks","N/A","~6",
   "Matches many micro FC patterns","Verify FC hole spec (M2 vs M2.5)","sunfounder kit","confirmed",
   "Some micro FCs are M2.5; have both on hand."),
 R(5,"Blue (removable) thread-locker","Medium-strength","N/A","~6",
   "Keeps steel motor screws from backing out","Don't use on nylon","oscarliang screws","confirmed",
   "Blue Loctite on STEEL motor screws only."),
]
# 17 microSD
microsd = [
 R(1,"Samsung PRO Endurance 128GB","V30/U3, ~100MB/s R, ~40MB/s W, 43,800h rated","N/A","~25",
   "Highest endurance rating; built for continuous write","Write speed modest vs Extreme","notebookcheck; gsmarena","confirmed",
   "Endurance class is the right tool for repetitive logging writes. RECOMMENDED."),
 R(2,"SanDisk High Endurance 128GB","V30/U3, up to 20,000h (256GB rated)","N/A","~22",
   "Widely available; proven dashcam endurance","Slightly lower rating than Samsung PRO","oempcworld","confirmed",
   "Strong ubiquitous alternative."),
 R(3,"SanDisk MAX Endurance 64GB","V30/U3, top SanDisk endurance tier","N/A","~25",
   "Top SanDisk endurance tier; long write life","Cost/GB higher","whiteboxstorage","confirmed",
   "If you want SanDisk's max write-life tier."),
 R(4,"Samsung PRO Endurance 64GB","V30/U3, 26,280h rated","N/A","~17",
   "Cheaper; ample for flight logs","Half the rated hours of 128GB","notebookcheck","confirmed",
   "Cost-effective if 64GB is enough per session."),
 R(5,"SanDisk Extreme (V30, non-endurance)","High burst write, 4K-rated","N/A","~15",
   "Fast for short high-bitrate logs","NOT endurance-rated — wears under continuous logging","oempcworld","confirmed",
   "Only for short/intermittent logs, not continuous."),
]
# 18 PROP GUARDS (safety-critical)
guards = [
 R(1,"Printed nylon (PA12/PA-CF) open guard ring","Stiff, high yield strain; tune wall to meet deflection SF","est. 8-14/set","filament",
   "Highest stiffness-to-mass => easiest to hit <=1.0mm deflection (SF>=2.0); lower thrust loss than full duct","Nylon needs dry filament/enclosure to print","uavmodel filament guide","inferred",
   "Stiffness satisfies the clearance-to-deflection SF, not just impact survival."),
 R(2,"Printed TPU (Shore 95A) open guard","Elastic, bounces off obstacles, won't shatter","est. 10-16/set","filament",
   "Best impact/crash survival; protects bystanders","Flexes more => must thicken to keep deflection<=1.0mm","uavmodel; intofpv 3d-guards","confirmed",
   "The safety material; deflection under load is the constraint to verify."),
 R(3,"BetaFPV Pavo20 Pro prop guard","Injection-molded guard sized for 2\" class","est. 6-10","6-10",
   "Validated geometry; consistent; cheap","Fits BetaFPV geometry, not arbitrary custom frame","getfpv.com pavo20-guard","confirmed",
   "Off-the-shelf clearance reference; still needs your SF deflection check."),
 R(4,"Commercial CineWhoop full duct (2\")","Closed duct; max protection + static-thrust gain","est. 15-30","10-20",
   "Best impact protection; static thrust gain from duct","Heaviest (may break 200g); tip-gap critical","getfpv.com guards-and-ducts","confirmed",
   "Only if mass budget allows; ducting changes thrust/efficiency + prop choice."),
 R(5,"Printed PETG open guard","Rigid, easy to print, decent stiffness","est. 8-14","filament",
   "Cheapest; prints on any machine; stiff","Brittle — breaks in 'the perfect crash'","uavmodel filament guide","confirmed",
   "OK for frame, weaker as the deflecting guard ring."),
 R(6,"Printed PETG-CF / PA-CF hybrid","Carbon-filled, very stiff, light","est. 7-12","filament",
   "High stiffness/mass => tightest deflection control","Brittle (PETG-CF); abrasive to nozzle","uavmodel filament guide","inferred",
   "If you want stiffness-led design and accept brittleness."),
 R(7,"TPU/nylon co-print (stiff ring + soft bumper)","Rigid nylon load ring + TPU contact edge","est. 12-18","filament",
   "Combines stiffness (meets SF) with impact compliance","Two-material printing complexity","intofpv 3d-guards","inferred",
   "BEST-of-both for a guarded UAV if you can dual-material print. Verify <=1.0mm deflection by FEA/push test."),
]
# 15 BENCH / TEST EQUIPMENT
bench = [
 R(1,"Tyto/RCbenchmark Series 1585","Thrust<=5kgf, torque<=2Nm, V, I, RPM (optical), temp, vibration, efficiency","N/A","625-845",
   "Full propulsion characterization incl torque->true efficiency + vibration (informs soft-mount)","Expensive; overkill if you only need thrust","tytorobotics 1580-1585","confirmed",
   "Only stand here giving torque + vibration — the data you want for 1103 prop matching. TOP for rigor."),
 R(2,"RCbenchmark Series 1520 (legacy/used)","Thrust<=5kgf +/-0.5%, V, I, RPM, efficiency (no torque)","N/A","~165 used",
   "Cheap if found used; proven thrust/efficiency sweeps","Discontinued; no torque/vibration","oscarliang 1520; tytorobotics","confirmed",
   "Best value if sourced used; covers thrust/current validation."),
 R(3,"DIY load cell (1-2kg) + HX711 + Arduino + INA219","Thrust + V/I (INA219) + RPM (optical tach); ~2-5g accuracy","N/A","15-40",
   "Cheapest; adequate for 1103 thrust; teaches the measurement chain","No torque; DIY calibration rigor on you","zbotic guide; cuds.soc.srcf","confirmed",
   "Downsize to 1-2kg cell (5kg wastes resolution on a ~400-600g-thrust machine)."),
 R(4,"Turnigy thrust stand","Basic thrust + current readout","N/A","40-70",
   "Plug-and-play cheap commercial","Coarse; limited logging","hobbyking","inferred",
   "Budget commercial middle ground."),
 R(5,"PVC + netting test cage","Enclosure for spin-up / abort testing","N/A","30-60",
   "Contains prop/guard failure during validation; cheap","Must be sized so guard fragments can't escape","community/DIY","inferred",
   "SAFETY-CRITICAL for a guarded UAV: run all thrust/guard tests inside it."),
 R(6,"Controlled-release / tethered fixture","Restrains airframe during powered tests; defined release","N/A","10-30",
   "Prevents fly-away during abort/runaway","DIY; verify restraint rating > max thrust","community/DIY","inferred",
   "Pair with the cage for any powered-prop test."),
 R(7,"Optical RPM probe (standalone)","Non-contact prop RPM","N/A","20-40",
   "Adds RPM to a DIY or thrust-only stand","Needs reflective marker / line of sight","tytorobotics accessories","confirmed",
   "Cheap upgrade to close the data gap on a DIY stand."),
]

SHEETS = [
 ("04_Motors", motors, "Motors (1103-class, 2S) — CRITICAL. Envelope: T/W>=2.0 needs >=100g/motor @200g AUW. EX1103 11000KV = 121.9g/motor (mfr, 2S+2023R)."),
 ("05_Propellers", props, "Propellers 2.0-2.3\" — CRITICAL. Gemfan 2023-3 is the prop behind the validated EX1103 thrust figure."),
 ("06_Battery", battery, "2S LiPo — CRITICAL. 4-motor peak = 36.8A. Continuous capability = C x Ah. 450mAh 75C = 33.75A FAILS; 550mAh 75C = 41.25A passes. Prefer XT30 over BT2.0 at this current."),
 ("01_FlightController", fc, "Flight controller — CRITICAL. KEY FORK: H7 (Kakute H7 Mini / Matek H743) for ArduPilot/PX4 autonomous recovery, OR AIO (Happymodel X12) for Betaflight stabilize-only + lowest mass."),
 ("02_ESC_AIO", esc, "ESC / 4-in-1 — CRITICAL. 1103@11000KV on 2S draws modest current; optimize MASS not amps. Avoid 35-50A boards."),
 ("06b_Vision_Compute", vision, "Vision/compute board — CRITICAL (autonomy). Release detection = motion fusion; Nicla bundles cam+IMU+ToF. NPU only on Coral/Sipeed/OAK. Confirm Coral mass."),
 ("07_Camera", camera, "Camera/image sensor. ROLLING shutter (OV5640/GC2145/OV2640) smears frames during a throw; GLOBAL shutter (OV9281/OV7251) is correct but needs a CSI host (Pi)."),
 ("11a_OpticalFlow", oflow, "Optical-flow positioning (indoor primary). Matek 3901-L0X bundles flow + lidar height in 2g, native ArduPilot/INAV."),
 ("11b_GPS", gps, "GPS (outdoor option only). Poor indoors + adds mass => optical flow is primary for the recovery use case."),
 ("08_ELRS_RX", elrs, "ELRS 2.4GHz receiver. On a LOS micro, mass+antenna placement matter more than range; diversity (RP3) only if the guard shadows antennas."),
 ("09_Antennas_Video", antennas, "Video antennas (control-link antenna ships with the ELRS RX). Match the connector (U.FL vs MMCX vs IPEX-1) to the VTX."),
 ("10a_FPV_Analog", fpv_analog, "FPV analog. ~2.6g / ~1.3W. NOTE: most nano VTX are 5V-input => need a 5V BEC, not raw 2S."),
 ("10b_FPV_Digital", fpv_digital, "FPV digital. VERDICT: only HDZero Whoop Lite fits sub-200g AND native 2S, but draws ~7W. DJI O4 / Walksnail Nano V3 not viable (power/mass)."),
 ("12_Buzzer_LED", buzzer, "Buzzer + status LED. Self-powered finder survives LiPo ejection; WS2812 adds orientation/arming status at ~0.5g."),
 ("13_Wiring_Connectors", wiring, "Wiring/connectors. Scheme: XT30 battery + 22AWG main / 26AWG motor leads + JST-SH signal + MMCX VTX antenna."),
 ("14_Charger", charger, "LiPo charger (support gear, non-flying). ToolkitRC M6AC = AC+DC in one box."),
 ("16_Fasteners", fasteners, "Fasteners/standoffs. Steel M2 for motors (blue threadlocker); nylon standoffs + silicone soft-mounts for the FC (protects the stabilization gyro)."),
 ("17_microSD", microsd, "microSD for onboard vision logging. Use ENDURANCE-class (continuous-write); min V30/U3."),
 ("18_PropGuards", guards, "Prop guards — SAFETY-CRITICAL. Requirement: >=2.0mm clearance, >=2.0 SF => max deflection <=1.0mm at the prop-facing edge. Stiffness-led (nylon) or stiff-ring+TPU co-print; VERIFY by FEA / static push test."),
 ("15_Bench_Test", bench, "Bench/test equipment (non-flying). Do ALL powered tests inside a netting cage + tethered release fixture."),
]
for title, rows, note in SHEETS:
    add_sheet(title, rows, note)

# ---- 00 Method & 01 Recommended Build ----------------------------------------
m = wb.active; m.title = "00_Method_Envelope"
lines = [
 ("MechAudit-style component selection — Guarded Micro-UAV (SelfStabilizingDrone)", True),
 ("Generated 2026-06-19 from 5 clustered deep-research streams vs primary/vendor sources.", False),
 ("", False),
 ("DESIGN ENVELOPE", True),
 ("FAA sub-250g; planning mass target <=200g; hard abort ceiling 225g.", False),
 ("Power 2S LiPo; motors 1103-class ~11000KV; props 2.0-2.3\"; static T/W >= 2.0.", False),
 ("Onboard vision/compute for release-event detection + self-stabilization recovery; ELRS 2.4GHz; printed PETG/nylon frame w/ prop guards.", False),
 ("", False),
 ("KEY ENGINEERING SCREENS", True),
 ("Thrust-to-weight: T/W>=2.0 @200g AUW => >=400g total => >=100g/motor. EX1103 11000KV+Gemfan 2023R = 121.9g/motor (mfr, 2S WOT) => 4x=487.6g => T/W 2.44 @200g (3.6 @134g).", False),
 ("Battery current screen: 4-motor peak = 4 x 9.2A = 36.8A. Continuous capability = C-rating x capacity(Ah). 450mAh 75C = 33.75A => FAILS (matches prior screening). 550mAh 75C = 41.25A => passes. Use XT30 (BT2.0 limited ~9-30A).", False),
 ("", False),
 ("CONFIDENCE LEGEND", True),
 ("confirmed (green) = from a cited primary/vendor source.  inferred (orange) = estimated/derived; verify before purchase.", False),
 ("", False),
 ("DECISIONS — Stage 1 status", True),
 ("RESOLVED: autonomy fork LOCKED to the H7/F7 RESEARCH path (ArduPilot/PX4 => logging, estimator access, failsafes) over the mass-minimal Betaflight AIO — the mass budget passes with wide margin, so recovery-research depth wins. See 01_Stage1_Locked_BOM.", False),
 ("STILL OPEN (deferred, do NOT block Stage 1): analog vs HDZero FPV (deferred to Stage 2 — not needed for cage hover); open guards vs full ducts (changes prop choice); optical-flow LOCKED for indoor Stage 1, GPS is an outdoor add-on.", False),
 ("", False),
 ("BIGGEST UNCONFIRMED ITEMS TO VERIFY ON A THRUST STAND / SCALE", True),
 ("- No INDEPENDENT bench for EX1103+2023 on 2S (121.9g is Happymodel's own). - Coral Dev Board Micro mass unpublished (decisive for the NPU option). - Printed guard mass + the <=1.0mm deflection SF must be measured, not assumed. - Real 2S/1103 peak current must be bench-measured before final ESC/battery sizing.", False),
]
for i,(txt,hd) in enumerate(lines,1):
    c=m.cell(i,1,txt); c.alignment=Alignment(wrap_text=True,vertical="top")
    if hd: c.font=Font(bold=True,color="1F3864",size=12)
m.column_dimensions["A"].width=140

rb = wb.create_sheet("01_Stage1_Locked_BOM")
rb.cell(1,1,"STAGE 1 LOCKED BOM — autonomy fork resolved to the H7/F7 RESEARCH path (ArduPilot/PX4: logging, estimator access, failsafes) over the mass-minimal Betaflight AIO. The mass budget passes with wide margin, so the recovery-research benefits win. FPV + GPS deferred to Stage 2.").font=Font(bold=True,color="1F3864")
rb.merge_cells(start_row=1,start_column=1,end_row=1,end_column=5)
rb.cell(1,1).alignment=WRAP_TOP; rb.row_dimensions[1].height=46
rb_cols = ["Subsystem","Locked part (Stage 1)","Nominal mass (g)","Price ~USD","Why this / rejected alternative"]
rb_rows = [
 ["Flight controller","Holybro Kakute H7 Mini (H743, 6 UART, baro)",5.5,65,"LOCKED H7 research path: ArduPilot/PX4, MAVLink companion link, blackbox logging, estimator access, failsafes. Rejected: Happymodel X12 AIO (lighter but Betaflight-only — no estimator/failsafe depth for recovery research)."],
 ["ESC (4-in-1)","HGLRC 13A 2-3S BB2 (20x20)",3.0,18,"20x20 matches the H7 FC mount; 13A/ch >> 9.2A/motor; 3g. Rejected: FD13A (16x16 — won't stack on 20x20); Tekko32 45-50A (over-spec, up to 15.6g). Upgrade: BLHeli_32/AM32 telemetry ESC if motor RPM/current logging wanted."],
 ["Motors x4","Happymodel EX1103 11000KV",15.2,36,"Only motor with a published 2S+2023 thrust curve (121.9g/motor => 4x=488g). Rejected: <10000KV (fails T/W); T-Motor F1103 (+~5g, thrust table unconfirmed)."],
 ["Propellers (+spares)","Gemfan Hurricane 2023-3",3.5,6,"The exact prop behind the validated thrust figure (removes inference risk). Rejected: 2.5\" pitch (more current draw); sub-2\" (less thrust)."],
 ["Battery","GNB 2S 550mAh 100C (XT30)",29.0,14,"55A cont >> 36.8A 4-motor peak. Rejected: 450mAh 75C (33.75A FAILS the current screen — your prior failure); BT2.0 (connector-limited). Budget alt: BetaFPV LAVA 550 75C (41.25A)."],
 ["Vision / compute","Arduino Nicla Vision (cam+IMU+ToF)",19.8,90,"Release detection = motion fusion; Nicla bundles camera+6-axis IMU+ToF in 20g, UART to FC. Rejected for Stage 1: Coral Micro (true NPU but mass unpublished); OAK-D-Lite (61g/5W)."],
 ["Positioning","Matek 3901-L0X optical flow + lidar",2.0,26,"Flow + height in 2g, native ArduPilot — pairs with the H7 path for indoor position/alt hold. GPS deferred (outdoor only)."],
 ["Control link RX","BetaFPV ELRS Lite (+ antenna)",0.46,12,"Lightest; the H7 FC has no onboard RX. Reliability upgrade: RadioMaster RP3 diversity (4.6g) if the guard shadows the antenna."],
 ["Buzzer + LED","VIFLY Finder Mini + WS2812 strip",3.3,18,"Self-powered finder survives LiPo ejection; WS2812 gives arming/orientation status (~0.5g)."],
 ["Wiring / connectors","XT30 + 22/26AWG silicone + JST-SH",5.5,8,"XT30 for 36.8A peak headroom (BT2.0 connector-limited). Weigh the finished harness."],
 ["Frame + guards + fasteners","Printed nylon frame + stiff guards + steel M2",42.5,40,"DESIGN WORK (your items 10/16/18). Guard ring must meet <=1.0mm deflection at the prop edge (SF>=2.0) — verify by FEA + static push test. Biggest single mass + the CAD-dependent unknown."],
 ["Vision logging","Samsung PRO Endurance 128GB microSD",0.4,25,"Endurance-class for continuous onboard logging (min V30/U3)."],
 ["FPV video","DEFERRED to Stage 2",0.0,0,"Not needed for manual cage hover (REQ-FLIGHT-001 is LOS). If added: analog Caddx Ant Lite + Reaper Nano (~3.6g, needs 5V BEC); HDZero Whoop Lite for HD (~6g but ~7W)."],
 ["GPS","DEFERRED (outdoor only)",0.0,0,"Indoor recovery uses optical flow. Outdoor add-on: Matek M10Q-5883 (8g)."],
]
total = sum(r[2] for r in rb_rows)
cost = sum(r[3] for r in rb_rows)
hdr = 2
for c,name in enumerate(rb_cols,1):
    cell=rb.cell(hdr,c,name); cell.fill=HEAD_FILL; cell.font=HEAD_FONT; cell.alignment=Alignment(horizontal="center",vertical="center"); cell.border=BORDER
for i,row in enumerate(rb_rows,hdr+1):
    for c,val in enumerate(row,1):
        cell=rb.cell(i,c,val); cell.alignment=WRAP_TOP; cell.border=BORDER
r=len(rb_rows)+hdr+1
rb.cell(r,1,"NOMINAL FLYING TOTAL (Stage 1, FPV/GPS deferred)").font=Font(bold=True)
rb.cell(r,3,round(total,1)).font=Font(bold=True)
rb.cell(r,4,cost).font=Font(bold=True)
rb.cell(r+1,1,f"vs 200g target: {round(200-total,1)}g margin   |   vs 225g abort ceiling: {round(225-total,1)}g margin   |   T/W @ this AUW = {round(487.6/total,2)}  (>=2.0 required)").font=Font(bold=True,color="375623")
rb.cell(r+2,1,"Frame/guard mass (42.5g nominal) dominates and is the CAD-dependent unknown — even at worst-case (frame 50g, vision 25g) + Stage-2 FPV (~3.6g) the build stays well under the 225g ceiling. Support gear (non-flying, not in total): ToolkitRC M6AC charger ~$60; thrust stand (DIY 1-2kg load cell ~$30 or used RCbenchmark 1520 ~$165); PVC+net cage ~$40; tethered release fixture ~$20.").alignment=WRAP_TOP
rb.merge_cells(start_row=r+2,start_column=1,end_row=r+2,end_column=5)
for col,w in zip("ABCDE",[20,40,16,12,82]):
    rb.column_dimensions[col].width=w

# order sheets: method, recommended, then categories
order = ["00_Method_Envelope","01_Stage1_Locked_BOM"] + [t for t,_,_ in SHEETS]
wb._sheets.sort(key=lambda s: order.index(s.title) if s.title in order else 99)
out = "Component_Selection_2026.xlsx"
wb.save(out)
print("wrote", out, "with", len(wb.sheetnames), "sheets:", wb.sheetnames)
